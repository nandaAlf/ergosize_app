from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status
from rest_framework.response import Response
from app1.models import Person, Measurement, Study, Dimension, StudyDimension
from app1.api.serializer import  PersonSerializer, MeasurementSerializer, StudyDimensionSerializer, StudySerializer, DimensionSerializer, StudyDetailWithPersonsSerializer
from django.http import Http404, HttpResponse, JsonResponse
from django.views import View
from django.db import connection
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from django.utils import timezone
from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
import io
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated
from accounts.api.permissions import HasRole, IsAdmin, IsAdminOrInvestigator, IsCreatorOrAdmin, IsInvestigator, IsInvestigatorOfPerson
# IsAdmin, IsCreator, IsGeneralUser, IsInvestigator
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from reportlab.lib.pagesizes import A4, landscape,letter
from rest_framework import pagination

class PersonViewSet(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    queryset = Person.objects.all()
    serializer_class = PersonSerializer
   
   
    def get_permissions(self):
        # list & retrieve
        if self.action in ['list', 'retrieve','update','partial_update','destroy']:
        #     # admin ve todo, investigador también pasa permiso pero
        #     # el queryset limita, y detail pasa a has_object_permission
            return [IsAuthenticated(),  IsInvestigatorOfPerson()]
        # create
        if self.action == 'create':
            return [IsAuthenticated(), IsAdminOrInvestigator()]
        # update / partial_update / destroy
        # if self.action in ['update','partial_update','destroy']:
        #     return [IsAuthenticated(), IsInvestigatorOfPerson()]
        # fallback: bloqueo
        return [IsAuthenticated(),IsInvestigatorOfPerson()]

    def get_queryset(self):
        user = self.request.user
        qs = Person.objects.all()
        if user.role == 'admin':
            return qs
        if user.role == 'investigador':
            # solo personas con mediciones en estudios supervisados
            return qs.filter(
                measurements__study__supervisor=user
            ).distinct()
        # usuarios generales no tienen acceso
        return Person.objects.none()

    def perform_create(self, serializer):
    
        user = self.request.user
        measurements = self.request.data.get('measurements', [])
        roles=['investigador','admin']
        if user.role in roles:
            if not measurements:
                raise PermissionDenied("Debes proporcionar al menos una medición.")

            # Extraer los IDs de estudios desde las mediciones
            study_ids = []
            for m in measurements:
                study_id = m.get('study_id')
                if not study_id:
                    raise PermissionDenied("Cada medición debe incluir un study_id.")
                study_ids.append(study_id)

            # Verificar que el usuario supervisa todos los estudios
            unauthorized = Study.objects.filter(
                id__in=study_ids
            ).exclude(supervisor=user)

            if unauthorized.exists():
                raise PermissionDenied("No puedes crear personas con mediciones de estudios que no supervisas.")

        serializer.save()
        
    @action(detail=True, methods=['get'], url_path='studies/(?P<study_id>[^/.]+)/measurements')
    def measurements_by_study(self, request, pk=None, study_id=None):
        """
        GET /api/persons/{person_id}/studies/{study_id}/measurements/
        Devuelve solo las mediciones de esa persona en ese estudio,
        agrupadas por categoría de dimensión.
        """
        # 1) Persona
        try:
            person = Person.objects.get(pk=pk)
        except Person.DoesNotExist:
            return Response({"detail": "Persona no encontrada."}, status=status.HTTP_404_NOT_FOUND)

        # 2) Mediciones de esa persona en el estudio
        qs = Measurement.objects.filter(person=person, study_id=study_id).select_related('dimension')

        # 3) Mapa de categoría
        labels = dict(Dimension.CATEGORY_CHOICES)

        grouped = {}
        for m in qs:
            dim = m.dimension
            cat = labels.get(dim.category, dim.category)
            grouped.setdefault(cat, []).append({
                'id_dimension': dim.id,
                'name':         dim.name,
                'initial':      dim.initial,
                'value':        m.value,
                'position':     m.position,
                'date':         m.date.isoformat(),
            })

        # 4) Datos de la persona (sin incluir mediciones planas)
        person_data = {
            'id':            person.id,
            'name':          person.name,
            'gender':        person.gender,
            'date_of_birth': person.date_of_birth.isoformat() if person.date_of_birth else None,
            'country':       person.country,
            'state':         person.state,
            'province':      person.province,
        }

        return Response({
            # 'person':       person_data,
            'id':            person.id,
            'name':          person.name,
            'gender':        person.gender,
            'date_of_birth': person.date_of_birth.isoformat() if person.date_of_birth else None,
            'country':       person.country,
            'state':         person.state,
            'province':      person.province,
            'measurements': grouped
        }, status=status.HTTP_200_OK)
   

    # def perform_update(self, serializer):
    #     """
    #     Igual que create: verifica que el investigador solo modifique
    #     mediciones en sus propios estudios.
    #     """
    #     user = self.request.user
    #     data = self.request.data.get('measurements', [])
    #     if user.role == 'investigador':
    #         study_ids = {m.get('study') for m in data}
           
    #         not_owned = Study.objects.filter(id__in=study_ids).exclude(supervisor=user)
    #         if not_owned.exists():
    #             raise PermissionDenied("No puedes modificar personas con mediciones de estudios que no supervisas.")
    #     serializer.save() 
   
class DimensionViewSet(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    queryset = Dimension.objects.all()
    serializer_class = DimensionSerializer
    
    def list(self, request, *args, **kwargs):
        """
        Sobrescribimos list() para devolver un dict { categoría: [dim1, dim2, ...], ... }
        """
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        data = serializer.data

        # Mapea el valor de choice ('0','1',...) a la etiqueta ("Altura",...)
        labels = dict(Dimension.CATEGORY_CHOICES)

        grouped = {}
        for item in data:
            cat_value = item['category']
            cat_label = labels.get(cat_value, cat_value)
            grouped.setdefault(cat_label, []).append(item)

        return Response(grouped)
    

class MeasurementViewSet(viewsets.ModelViewSet):
    queryset = Measurement.objects.all()
    authentication_classes = [JWTAuthentication]
    serializer_class = MeasurementSerializer
    # permission_classes = [IsAuthenticated]  # Solo usuarios autenticados pueden eliminar

    # def get_permissions(self):
    #     if self.action in ['create', 'update', 'partial_update', 'destroy']:
    #         return [IsCreatorOrAdmin()]
    #     return [IsAuthenticated()]
    
    def get_permissions(self):
        # Para destruir mediciones en bloque, comprobamos permiso sobre el estudio
        if self.action == 'destroy':
            return [IsAuthenticated(), IsCreatorOrAdmin()]
        # Resto de acciones: cualquier usuario autenticado
        return [IsAuthenticated()]

    def destroy(self, request, *args, **kwargs):
        """
        DELETE /api/measurements/study_id/?&person_id=Y
        Borra todas las mediciones de esa persona en ese estudio,
        y si la persona no tiene más mediciones en otros estudios,
        elimina también el registro de Person.
        """
        # study_id  = request.query_params.get('study_id')
        study_id=self.kwargs.get('pk')
        person_id = request.query_params.get('person_id')

        if not study_id or not person_id:
            return Response(
                {"error": "Se requieren 'study_id' y 'person_id' como query params."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar permiso: simulamos un objeto Study para IsCreatorOrAdmin
        # from studies.models import Study
        try:
            study = Study.objects.get(id=study_id)
        except Study.DoesNotExist:
            return Response(
                {"error": "Estudio no encontrado."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Chequear permiso de objeto
        self.check_object_permissions(request, study)

        # Borrar mediciones en ese estudio
        deleted_count, _ = Measurement.objects.filter(
            study_id=study_id,
            person_id=person_id
        ).delete()

        if deleted_count == 0:
            return Response(
                {"error": "No se encontraron mediciones para ese estudio y persona."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Si la persona ya no tiene mediciones en ningún estudio, borrarla
        remaining = Measurement.objects.filter(person_id=person_id).exists()
        if not remaining:
            try:
                person = Person.objects.get(id=person_id)
                person.delete()
            except Person.DoesNotExist:
                pass  # ya puede no existir

        return Response(
            {"message": "Mediciones eliminadas correctamente."},
            status=status.HTTP_204_NO_CONTENT
        )
class StudyViewSet(viewsets.ModelViewSet):

    authentication_classes = [JWTAuthentication]
    queryset = Study.objects.all()
    serializer_class = StudySerializer
    permission_classes = [IsAuthenticated]
    pagination_class = pagination.PageNumberPagination
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    # def get_queryset(self):   
    #     # Ordenar por fecha de inicio descendente y optimizar consultas
    #     qs = (
    #         super().get_queryset()
    #         .select_related('supervisor')  # Optimizar relaciones
    #         .only(  # Seleccionar solo campos necesarios
    #             'id', 'name', 'size', 'location', 'country', 
    #             'start_date', 'end_date',
    #             'supervisor__username'  # Solo username del supervisor
    #         )
    #         .order_by('-start_date')  # Orden descendente por fecha
    #     )
        
    #     # Si se pasa ?mine=true, filtra por supervisor
    #     if self.request.query_params.get('mine') == 'true':
    #         return qs.filter(supervisor=self.request.user)
            
    #     return qs
    def get_queryset(self):   
        qs = super().get_queryset()
        # Búsqueda
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(name__icontains=search) | 
                Q(location__icontains=search) | 
                Q(country__icontains=search)
            )
        
        # Filtro por género
        gender = self.request.query_params.get('gender')
        if gender:
            qs = qs.filter(gender=gender)
        
        # Rango de fechas
        start_date_gte = self.request.query_params.get('start_date__gte')
        if start_date_gte:
            qs = qs.filter(start_date__gte=start_date_gte)
            
        start_date_lte = self.request.query_params.get('start_date__lte')
        if start_date_lte:
            qs = qs.filter(start_date__lte=start_date_lte)
        # Obtener parámetro de ordenamiento (default: -start_date)
        ordering = self.request.query_params.get('ordering', '-start_date')
         # Validar campos permitidos para ordenamiento
        valid_fields = {'start_date', 'end_date', 'name', 'size'}
        if ordering.lstrip('-') in valid_fields:
            qs = qs.order_by(ordering)
        else:
            # Orden por defecto si el campo no es válido
            qs = qs.order_by('-start_date')
        # Si se pasa ?mine=true, filtra por supervisor
        if self.request.query_params.get('mine') == 'true':
            return qs.filter(supervisor=self.request.user)
        return qs
    
    def get_permissions(self):
        if   self.action == 'create':
            # Solo admin e investigadores pueden crear
            return [IsAdminOrInvestigator()]
        elif self.action in ['update', 'partial_update', 'destroy']:
            # Solo el creador (supervisor) o admin pueden modificar o eliminar
            return [IsCreatorOrAdmin()]
        elif self.action == 'members':
            # Solo el creador o admin pueden ver las personas y sus mediciones
            return [IsCreatorOrAdmin()]
        # Listar y recuperar: cualquier usuario autenticado
        return [IsAuthenticated()]
    
    @action(detail=True, methods=['get'], url_path='members')
    def members(self, request, pk=None):
        """
        GET /api/studies/{pk}/members/
        devuelve el estudio con la lista de personas y sus valores de medición.
        """
        study = self.get_object()
        serializer = StudyDetailWithPersonsSerializer(study, context={'request': request})
        return Response(serializer.data)

class StudyDimensionViewSet(viewsets.ModelViewSet):
    queryset = StudyDimension.objects.all()
    serializer_class = StudyDimensionSerializer
    
    # def get_permissions(self):
    #     if self.action in ['create', 'update', 'partial_update', 'destroy']:
    #         # sólo admin e investigadores
    #         permission_classes = [IsAdmin|IsInvestigator]
    #     else:
    #         # listar y retrieve permitidos a todos autenticados
    #         permission_classes = [IsAdmin|IsInvestigator|IsGeneralUser]
    #     return [perm() for perm in permission_classes]




class Perceptil(View):
    """Vista para obtener percentiles de un estudio."""
    def get(self, request, study_id):
        if not study_id:
            return JsonResponse(
                {"status": "error", "message": "El parámetro 'study_id' es requerido."},
                status=400
            )

        # Parámetros GET
        gender = request.GET.get('gender')           # 'M', 'F' o 'mixto'
        age_ranges_param = request.GET.get('age_ranges')   # e.g. "10-12,12-14,14-16"
        dimensions = request.GET.get('dimensions')   # e.g. "1,2,3 ,..id de las dimensiones"
        percentiles_param = request.GET.get('percentiles') # e.g. "5,95"

        # Validaciones básicas
        if gender and gender not in ('M', 'F', 'mixto'):
            return JsonResponse({'status':'error','message':'Valor gender inválido.'}, status=400)

        # Parsear age_ranges
        try:
            age_ranges = [
                tuple(map(int, r.split('-')))
                for r in age_ranges_param.split(',')
            ] if age_ranges_param else None
        except Exception:
            return JsonResponse(
                {'status':'error','message':"'age_ranges' debe ser lista de rangos 'min-max'."},
                status=400
            )

        # Parsear dimensions
        try:
            dims_filter = [int(d) for d in dimensions.split(',')] if dimensions else None
        except ValueError:
            return JsonResponse(
                {'status':'error','message':"'dimensions' debe ser lista de IDs enteros."},
                status=400
            )

        # Parsear percentiles
        try:
            perc_list = [float(p) for p in percentiles_param.split(',')] if percentiles_param else None
        except ValueError:
            return JsonResponse(
                {'status': 'error', 'message': "'percentiles' debe ser lista de números."},
                status=400
            )

        # Llamada a la lógica de cálculo
        try:
            results = calculate_percentiles_for_study(
                study_id=study_id,
                gender=gender,
                age_ranges=age_ranges,
                dimensions_filter=dims_filter,
                percentiles_list=perc_list
            )
            return JsonResponse({'status': 'success', 'study_id': study_id, 'results': results})
        except Http404 as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


# def export_excel_percentiles(request, study_id):
#     try:
#         study = Study.objects.get(id=study_id)
#         name_study = request.GET.get('name', f"Estudio {study_id}")
#         gender = request.GET.get('gender')
#         age_min = int(request.GET.get('age_min')) if request.GET.get('age_min') else None
#         age_max = int(request.GET.get('age_max')) if request.GET.get('age_max') else None
#         dimensions_filter = request.GET.get('dimensions')
#         if dimensions_filter:
#             dimensions_filter = [int(x) for x in dimensions_filter.split(',')]

#         percentiles = request.GET.get('percentiles')
#         if percentiles:
#             percentiles_list = [float(p) for p in percentiles.split(',')]
#         else:
#             percentiles_list = [5, 10, 25, 50, 75, 90, 95]

#         # Construir age_ranges correctamente
#         age_ranges = [(age_min, age_max)] if age_min or age_max else None

#         results = calculate_percentiles_for_study(
#             study_id,
#             gender=gender,
#             age_ranges=age_ranges,
#             dimensions_filter=dimensions_filter,
#             percentiles_list=percentiles_list
#         )

#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Percentiles"

#         # --- Estilos
#         header_font = Font(bold=True)
#         title_font = Font(size=14, bold=True)
#         center_align = Alignment(horizontal="center")
#         left_align = Alignment(horizontal="left")
#         fill_gray = PatternFill("solid", fgColor="DDDDDD")
#         border_thin = Border(
#             left=Side(style='thin'), right=Side(style='thin'),
#             top=Side(style='thin'), bottom=Side(style='thin')
#         )

#         # --- Encabezado principal
#         ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5 + len(percentiles_list))
#         title_cell = ws.cell(row=1, column=1)
#         title_cell.value = f"Tabla antropométrica: {name_study}"
#         title_cell.font = title_font
#         title_cell.alignment = center_align
#         title_cell.fill = fill_gray

#         # --- Metadatos del estudio
#         metadata = [
#             ("Fecha inicio:", study.start_date.strftime('%Y-%m-%d') if study.start_date else "N/A"),
#             ("Fecha fin:", study.end_date.strftime('%Y-%m-%d') if study.end_date else "N/A"),
#             ("Tamaño muestra:", study.size),
#             ("Género:", {"M": "Masculino", "F": "Femenino", "mixto": "Mixto"}.get(gender, "Todos")),
#             ("Rango edad:", f"{age_min or 'N/A'} - {age_max or 'N/A'}")
#         ]

#         for i, (label, value) in enumerate(metadata, start=2):
#             ws.append([label, value])
#             for col in [1, 2]:
#                 cell = ws.cell(row=i, column=col)
#                 cell.font = Font(bold=(col == 1))
#                 cell.alignment = left_align

#         # --- Encabezados de datos
#         headers = ["Dimensión", "Género", "Rango Edad", "Media", "SD"] + [f"P{p}" for p in percentiles_list]
#         ws.append([])
#         ws.append(headers)
        
#         # Aplicar estilos a encabezados
#         for col in range(1, len(headers) + 1):
#             cell = ws.cell(row=ws.max_row, column=col)
#             cell.font = header_font
#             cell.alignment = center_align
#             cell.fill = fill_gray
#             cell.border = border_thin

#         # --- Datos principales
#         for res in results:
#             dim_name = res['dimension']
#             for gender_key, age_data in res.get('by_gender', {}).items():
#                 for age_range, stats in age_data.items():
#                     # Formatear valores
#                     gender_display = {
#                         'M': 'Masculino',
#                         'F': 'Femenino',
#                         'all': 'Mixto'
#                     }.get(gender_key, gender_key)
                    
#                     age_display = "Todos" if age_range == "all" else age_range
                    
#                     row = [
#                         dim_name,
#                         gender_display,
#                         age_display,
#                         round(stats['mean'], 2) if stats['mean'] else "-",
#                         round(stats['sd'], 2) if stats['sd'] else "-"
#                     ] + [round(stats['percentiles'].get(str(p), 0), 2) for p in percentiles_list]
                    
#                     ws.append(row)
                    
#                     # Formatear números
#                     for col in range(4, len(row) + 1):
#                         cell = ws.cell(row=ws.max_row, column=col)
#                         cell.number_format = '0.00'

#         # --- Ajustar anchos de columnas
#         column_widths = [35, 15, 15, 10, 10] + [10] * len(percentiles_list)
#         for i, width in enumerate(column_widths, start=1):
#             ws.column_dimensions[get_column_letter(i)].width = width

#         # --- Preparar respuesta
#         response = HttpResponse(
#             content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             headers={'Content-Disposition': f'attachment; filename="tabla_antropometrica_{study_id}.xlsx"'},
#         )
#         wb.save(response)
#         return response

#     except Study.DoesNotExist:
#         return HttpResponse("Estudio no encontrado.", status=404)
#     except Exception as e:
#         return HttpResponse(f"Error al generar el reporte: {str(e)}", status=500)
    


# def export_excel_percentiles(request, study_id):
#     try:
#         study = Study.objects.get(id=study_id)
#         name_study = request.GET.get('name', f"Estudio {study_id}")
#         gender = request.GET.get('gender')
#         age_min = int(request.GET.get('age_min')) if request.GET.get('age_min') else None
#         age_max = int(request.GET.get('age_max')) if request.GET.get('age_max') else None
#         dimensions_filter = request.GET.get('dimensions')
#         if dimensions_filter:
#             dimensions_filter = [int(x) for x in dimensions_filter.split(',')]

#         percentiles = request.GET.get('percentiles')
#         if percentiles:
#             percentiles_list = [float(p) for p in percentiles.split(',')]
#         else:
#             percentiles_list = [5, 10, 25, 50, 75, 90, 95]

#         age_ranges = [(age_min, age_max)] if age_min or age_max else None
#         results = calculate_percentiles_for_study(
#             study_id,
#             gender=gender,
#             age_ranges=age_ranges,
#             dimensions_filter=dimensions_filter,
#             percentiles_list=percentiles_list
#         )

#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Percentiles"

#         # Styles
#         header_font = Font(bold=True)
#         title_font = Font(size=14, bold=True)
#         center = Alignment(horizontal="center")
#         left = Alignment(horizontal="left")
#         gray_fill = PatternFill("solid", fgColor="DDDDDD")
#         thin_border = Border(
#             left=Side(style='thin'), right=Side(style='thin'),
#             top=Side(style='thin'), bottom=Side(style='thin')
#         )

#         # Title
#         ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(percentiles_list))
#         cell = ws.cell(row=1, column=1)
#         cell.value = f"Tabla antropométrica: {name_study}"
#         cell.font = title_font
#         cell.alignment = center
#         cell.fill = gray_fill

#         # Metadata
#         meta = [
#             ("Fecha inicio:", study.start_date.strftime('%Y-%m-%d') if study.start_date else "N/A"),
#             ("Fecha fin:", study.end_date.strftime('%Y-%m-%d') if study.end_date else "N/A"),
#             ("Tamaño muestra:", study.size),
#             ("Género filt:", { 'M': 'Masculino', 'F': 'Femenino', 'mixto': 'Mixto' }.get(gender, 'Todos')),
#             ("Edad filt:", f"{age_min or 'N/A'} - {age_max or 'N/A'}")
#         ]
#         for i, (lbl, val) in enumerate(meta, start=2):
#             ws.append([lbl, val])
#             for col in (1, 2):
#                 c = ws.cell(row=i, column=col)
#                 c.font = Font(bold=(col == 1))
#                 c.alignment = left

#         # Data header
#         ws.append([])
#         hdr = ["Dimensión"] + [f"P{int(p)}" for p in percentiles_list]
#         ws.append(hdr)
#         row_hdr = ws.max_row
#         for col in range(1, len(hdr) + 1):
#             c = ws.cell(row=row_hdr, column=col)
#             c.font = header_font
#             c.alignment = center
#             c.fill = gray_fill
#             c.border = thin_border

#         # Grouped data by sexo and edad
#         # Collect nested structure
#         grouped = {}
#         for entry in results:
#             dim = entry['dimension']
#             for g_key, ages in entry['by_gender'].items():
#                 grouped.setdefault(g_key, {})
#                 for age_key, stats in ages.items():
#                     grouped[g_key].setdefault(age_key, []).append((dim, stats))

#         # Write groups
#         for g_key, age_dict in grouped.items():
#             gender_disp = {'M': 'Masculino', 'F': 'Femenino', 'all': 'Mixto'}.get(g_key, g_key)
#             # Sexo header
#             ws.append([])
#             ws.append([f"Sexo: {gender_disp}"])
#             ws.merge_cells(start_row=ws.max_row, start_column=1,
#                            end_row=ws.max_row, end_column=1 + len(percentiles_list))
#             g_cell = ws.cell(row=ws.max_row, column=1)
#             g_cell.font = header_font
#             g_cell.fill = gray_fill
#             g_cell.alignment = left

#             for age_key, dims in age_dict.items():
#                 age_disp = age_key if age_key != 'all' else 'Todos'
#                 # Edad header
#                 ws.append([f"  Edad: {age_disp}"])
#                 ws.merge_cells(start_row=ws.max_row, start_column=1,
#                                end_row=ws.max_row, end_column=1 + len(percentiles_list))
#                 a_cell = ws.cell(row=ws.max_row, column=1)
#                 a_cell.font = header_font
#                 a_cell.fill = gray_fill
#                 a_cell.alignment = left

#                 # Dimensions rows
#                 for dim_name, stats in dims:
#                     row = [f"    {dim_name}"] + [round(stats['percentiles'].get(str(int(p)), 0), 2)
#                                                    for p in percentiles_list]
#                     ws.append(row)
#                     # Numeric format
#                     for col in range(2, len(row) + 1):
#                         ws.cell(row=ws.max_row, column=col).number_format = '0.00'

#         # Adjust widths
#         widths = [40] + [12] * len(percentiles_list)
#         for i, w in enumerate(widths, start=1):
#             ws.column_dimensions[get_column_letter(i)].width = w

#         # Response
#         response = HttpResponse(
#             content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             headers={'Content-Disposition': f'attachment; filename="tabla_antropometrica_{study_id}.xlsx"'},
#         )
#         wb.save(response)
#         return response

#     except Study.DoesNotExist:
#         return HttpResponse("Estudio no encontrado.", status=404)
#     except Exception as e:
#         return HttpResponse(f"Error al generar el reporte: {str(e)}", status=500)
    

def export_excel_percentiles(request, study_id):
    try:
        # --- Fetch study & params
        study = Study.objects.get(id=study_id)
        name_study = request.GET.get('name', f"Estudio {study.name}")
        gender = request.GET.get('gender')  # 'M','F','mixto' or None
        age_ranges_param = request.GET.get('age_ranges')  # "10-20,20-30"
        dimensions_filter = request.GET.get('dimensions')
        percentiles_param = request.GET.get('percentiles')

        # Parse filters
        age_ranges = [tuple(map(int,r.split('-'))) for r in age_ranges_param.split(',')] if age_ranges_param else [(None,None)]
        dims = [int(x) for x in dimensions_filter.split(',')] if dimensions_filter else None 
        perc_list = [float(p) for p in percentiles_param.split(',')] if percentiles_param else [5,95]

        results = calculate_percentiles_for_study(
            study_id=study_id,
            gender=gender,
            age_ranges=age_ranges,
            dimensions_filter=dims,
            percentiles_list=perc_list
        )

        # Preparar libro
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabla"

        # --- Estilos básicos
        title_font = Font(size=14, bold=True)
        hdr_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        # gray = PatternFill("solid", fgColor="DDDDDD")
        thin = Border(*[Side(style='thin')]*4)

        # --- Título y metadata
        ws.merge_cells("A1:Z1")
        c = ws['A1']; c.value = f"Tabla antropométrica: {name_study}"; c.font=title_font; 
        # c.alignment=center; c.fill=gray
        meta = [
            ("Fecha inicio", study.start_date),
            ("Fecha fin", study.end_date),
            ("Muestra", study.size),
            ("Género", {'M':'Hombres','F':'Mujeres','mixto':'Mixto'}.get(gender,'Todos')),
            ("Edad", ",".join(f"{a}-{b}" for a,b in age_ranges))
        ]
        for i,(k,v) in enumerate(meta, start=2):
            ws[f"A{i}"] = k; ws[f"B{i}"] = v if v else 'N/A'
            ws[f"A{i}"].font = hdr_font

        # --- Preparar encabezados multi-nivel
        sexes = []
        if gender=='mixto': sexes = ['M','F']
        elif gender in ('M','F'): sexes=[gender]
        else: sexes=['M','F']
        metrics = ['Media','SD'] + [f"P{int(p)}" for p in perc_list]
        cols_per_range = len(metrics)
        start = len(meta) + 3

        # Row1: Dimensión + Sexos
        ws.merge_cells(start_row=start, start_column=1, end_row=start+2, end_column=1)
        dim_cell=ws.cell(row=start, column=1, value="Dimensión").font=hdr_font; ws.cell(row=start, column=1).alignment=center; ws.cell(row=start, column=1).border=thin;
        # dim_cell.font = hdr_font
        # dim_cell.alignment = center
        # dim_cell.border = thin
        # border_cell=ws.cell(row=5,column=4)
        # border_cell.border=thin
        col_idx = 2
        for s in sexes:
            label = 'Hombres' if s=='M' else 'Mujeres'
            span = len(age_ranges) * cols_per_range
            ws.merge_cells(start_row=start, start_column=col_idx, end_row=start, end_column=col_idx+span-1)
            cell = ws.cell(row=start, column=col_idx, value=label)
            cell.font=hdr_font; cell.alignment=center; cell.border=thin
            # Row2: rangos
            for ar in age_ranges:
                age_label = f"{ar[0]}-{ar[1]}"
                ws.merge_cells(start_row=start+1, start_column=col_idx, end_row=start+1, end_column=col_idx+cols_per_range-1)
                c2 = ws.cell(row=start+1, column=col_idx, value=age_label)
                c2.font=hdr_font; c2.alignment=center;  c2.border=thin
                # Row3: metrics
                for m in metrics:
                    c3 = ws.cell(row=start+2, column=col_idx, value=m)
                    c3.font=hdr_font; c3.alignment=center;  c3.border=thin
                    col_idx += 1
        # --- Rellenar datos
        row = start+3
        for entry in results:
            ws.cell(row=row, column=1, value=entry['dimension'])
            col = 2
            for s in sexes:
                key_s = s
                for ar in age_ranges:
                    key_range = f"{ar[0]}-{ar[1]}"
                    stats = entry.get('by_gender',{}).get(key_s,{}).get(key_range,{})
                    # media, sd
                    ws.cell(row=row, column=col, value=stats.get('mean')); col+=1
                    ws.cell(row=row, column=col, value=stats.get('sd')); col+=1
                    # percentiles
                    for p in perc_list:
                        ws.cell(row=row, column=col, value=stats.get('percentiles',{}).get(str(int(p)),"")); col+=1
            row +=1

        # Ajustes de ancho
        for i in range(1, col):
            ws.column_dimensions[get_column_letter(i)].width = 15

        #bordes
        # border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # end_row = start + 3 + len(results) - 1  # fila final
        # end_col = 1 + len(sexes) * len(age_ranges) * len(metrics)  # columna final

        # for r in range(start, end_row + 1):
        #     for c in range(1, end_col + 1):
        #         cell = ws.cell(row=r, column=c)
        #         cell.border = thin

        # Enviar
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={'Content-Disposition': f'attachment; filename="tabla_{study_id}.xlsx"'}
        )
        wb.save(resp)
        return resp
    except Study.DoesNotExist:
        return HttpResponse(status=404, content="Estudio no encontrado")
    except Exception as e:
        return HttpResponse(status=500, content=f"Error: {e}")

# @csrf_exempt
# def preview_excel_percentiles(request):
#     print(request.method)
#     print(request.FILES)
#     print("Archivo recibido:", request.FILES.get("archivo"))
#     if request.method == "POST" and request.FILES.get("archivo"):
#         archivo = request.FILES["archivo"]
#         wb = load_workbook(archivo)
#         ws = wb["Percentiles"]

#         data = []
#         headers = []
#         for i, row in enumerate(ws.iter_rows(min_row=8, values_only=True), start=8):
#             if i == 8:
#                 headers = row
#                 continue
#             if not any(row):
#                 continue

#             item = {
#                 "dimension": row[0],
#                 "mean": row[1],
#                 "sd": row[2],
#                 "percentiles": {}
#             }
#             for j, value in enumerate(row[3:], start=3):
#                 if headers[j]:
#                     key = str(headers[j]).replace("P", "")
#                     item["percentiles"][key] = value
#             data.append(item)

#         return JsonResponse(data, safe=False)

#     return JsonResponse({"error": "Archivo no proporcionado o método incorrecto"}, status=400)

# def export_pdf_percentiles(request, study_id):
#     try:
#         # Obtener parámetros y estudio de igual forma que en Excel
#         study = Study.objects.get(id=study_id)
#         name_study = request.GET.get('name', f"Estudio {study_id}")
#         gender = request.GET.get('gender')
#         age_min = request.GET.get('age_min')
#         age_max = request.GET.get('age_max')
#         age_min = int(age_min) if age_min else None
#         age_max = int(age_max) if age_max else None

#         dimensions_filter = request.GET.get('dimensions')
#         dimensions_filter = [int(x) for x in dimensions_filter.split(',')] if dimensions_filter else None

#         percentiles = request.GET.get('percentiles')
#         percentiles_list = [float(p) for p in percentiles.split(',')] if percentiles else [5, 10, 25, 50, 75, 90, 95]

#         results = calculate_percentiles_for_study(
#             study_id,
#             gender=gender,
#             age_min=age_min,
#             age_max=age_max,
#             dimensions_filter=dimensions_filter,
#             percentiles_list=percentiles_list
#         )

#         # Preparar el documento PDF en memoria
#         buffer = BytesIO()
#         doc = SimpleDocTemplate(buffer, pagesize=A4, title=f"Tabla antropométrica: {name_study}")

#         styles = getSampleStyleSheet()
#         # Estilos personalizados
#         title_style = styles["Title"]
#         title_style.alignment = 1  # Centrado

#         header_style = styles["Heading4"]
#         normal_style = styles["Normal"]
#         normal_style.spaceAfter = 8

#         elements = []

#         # Título
#         elements.append(Paragraph(f"Tabla antropométrica: {name_study}", title_style))
#         elements.append(Spacer(1, 12))

#         # Metadatos (usamos una tabla para organizarlos, de 4 columnas)
#         meta_data = [
#             ["Fecha inicio:", str(study.start_date), "Fecha fin:", str(study.end_date)],
#             ["Muestra:", str(study.size), "Género:", gender if gender else "Mixto"],
#             ["Edad mínima:", str(age_min) if age_min is not None else "—", "Edad máxima:", str(age_max) if age_max is not None else "—"],
#         ]
#         meta_table = Table(meta_data, hAlign="LEFT", colWidths=[80, 120, 80, 120])
#         meta_table.setStyle(TableStyle([
#             ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
#             ('FONTSIZE', (0,0), (-1,-1), 10),
#             ('BOTTOMPADDING', (0,0), (-1,-1), 4),
#             ('TOPPADDING', (0,0), (-1,-1), 4),
#         ]))
#         elements.append(meta_table)
#         elements.append(Spacer(1, 12))

#         # Descripción: cabecera y texto en un párrafo aparte
#         elements.append(Paragraph("Descripción:", header_style))
#         elements.append(Paragraph(study.description, normal_style))
#         elements.append(Spacer(1, 12))

#         # Preparar datos para la tabla de resultados
#         # Encabezados
#         headers = ["Dimensión", "Media", "SD"] + [f"P{p}" for p in percentiles_list]
#         data_table = [headers]
        
#         # Agregar cada fila de resultados
#         for res in results:
#             row = [
#                 res["dimension"],
#                 round(res["mean"], 2),
#                 round(res["sd"], 2),
#             ]
#             row += [res["percentiles"].get(str(p), '') for p in percentiles_list]
#             data_table.append(row)
        
#         # Crear tabla con estilos
#         table = Table(data_table, hAlign="CENTER")
#         table.setStyle(TableStyle([
#             ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
#             ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
#             ('ALIGN', (0,0), (-1,-1), 'CENTER'),
#             ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
#             ('FONTSIZE', (0,0), (-1,0), 10),
#             ('BOTTOMPADDING', (0,0), (-1,0), 8),
#             ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
#         ]))
#         elements.append(table)
#         elements.append(Spacer(1, 12))

#         # Construir el documento
#         doc.build(elements)

#         buffer.seek(0)
#         response = HttpResponse(buffer, content_type='application/pdf')
#         response['Content-Disposition'] = f'attachment; filename="percentiles_study_{study_id}.pdf"'
#         return response

#     except Study.DoesNotExist:
#         return HttpResponse("Estudio no encontrado.", status=404)
#     except Exception as e:
#         return HttpResponse(f"Ocurrió un error: {str(e)}", status=500)
def export_pdf_percentiles(request, study_id):
    try:
         # Parámetros y estudio igual que en Excel
        study = Study.objects.get(id=study_id)
        name_study = request.GET.get('name', f"Estudio {study_id}")
        gender = request.GET.get('gender')
        age_ranges_param = request.GET.get('age_ranges')
        age_ranges = [tuple(map(int, r.split('-'))) for r in age_ranges_param.split(',')] if age_ranges_param else [(None, None)]
        dims_filter = request.GET.get('dimensions')
        dims = [int(x) for x in dims_filter.split(',')] if dims_filter else None
        perc_param = request.GET.get('percentiles')
        perc_list = [int(p) for p in perc_param.split(',')] if perc_param else [5, 95]

        results = calculate_percentiles_for_study(
            study_id=study_id,
            gender=gender,
            age_ranges=age_ranges,
            dimensions_filter=dims,
            percentiles_list=[float(p) for p in perc_list]
        )

        # Preparar PDF horizontal
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36) 
        styles = getSampleStyleSheet()
        elements = []

        # Título
        title = Paragraph(f"<b>Tabla antropométrica: {name_study}</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Metadata: asegurar 4 columnas
        meta = [
            ['Fecha inicio', study.start_date or '—', 'Fecha fin', study.end_date or '—'],
            ['Muestra', study.size, 'Género', {'M':'Hombres','F':'Mujeres','mixto':'Mixto'}.get(gender,'Mixto')],
            ['Edad ranges', ','.join(f"{a}-{b}" for a, b in age_ranges), '', '']
        ]
        table_meta = Table(meta, colWidths=[80, 120, 80, 120])
        table_meta.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(table_meta)
        elements.append(Spacer(1, 12))

        # Construir encabezados multinivel
        sexes = ['M', 'F'] if not gender or gender == 'mixto' else [gender]
        metrics = ['Media', 'SD'] + [f"P{p}" for p in perc_list]
        cols_per_range = len(metrics)

        # Primera fila (sexos)
        header_row1 = ['Dimensión']
        for s in sexes:
            label = {'M':'Hombres','F':'Mujeres'}.get(s, 'Mixto')
            header_row1 += [label] * (len(age_ranges) * cols_per_range)

        # Segunda fila (edades)
        header_row2 = ['']
        for _ in sexes:
            for (a, b) in age_ranges:
                age_label = f"{a}-{b}"
                header_row2 += [age_label] * cols_per_range

        # Tercera fila (metrics)
        header_row3 = [''] + metrics * len(age_ranges) * len(sexes)

        data = [header_row1, header_row2, header_row3]

        # Filas de datos
        for res in results:
            row = [res['dimension']]
            for s in sexes:
                for ar in age_ranges:
                    key = f"{ar[0]}-{ar[1]}"
                    stats = res.get('by_gender', {}).get(s, {}).get(key, {})
                    row.append(round(stats.get('mean', 0), 2))
                    row.append(round(stats.get('sd', 0), 2))
                    for p in perc_list:
                        row.append(round(stats.get('percentiles', {}).get(str(p), 0), 2))
            data.append(row)

                # Crear tabla con estilo completo de grid
        num_cols = 1 + len(sexes) * len(age_ranges) * cols_per_range
        # Calculamos colWidths para ocupar todo el ancho utilizable\ n       
        page_width, _ = landscape(letter)
        usable_width = page_width - doc.leftMargin - doc.rightMargin
        num_cols = len(data[0])
        first_col_width = usable_width * 0.3
        other_width = (usable_width - first_col_width) / (num_cols - 1)
        colWidths = [first_col_width] + [other_width] * (num_cols - 1)
        table = Table(data, colWidths=colWidths, repeatRows=3)
        # table = Table(data, repeatRows=3)
        style_commands = [
            # Span título 'Dimensión' vertical
            ('SPAN', (0, 0), (0, 2)),
            # Span por sexo en fila 0
        ]
        # Agregar spans dinámicos para Hombres/Mujeres
        col = 1
        for s in sexes:
            span_start = col
            span_end = col + len(age_ranges) * cols_per_range - 1
            style_commands.append(('SPAN', (span_start, 0), (span_end, 0)))
            col = span_end + 1
        # Span rangos de edad en fila 1
        col = 1
        for _ in sexes:
            for _ in age_ranges:
                age_span_start = col
                age_span_end = col + cols_per_range - 1
                style_commands.append(('SPAN', (age_span_start, 1), (age_span_end, 1)))
                col = age_span_end + 1
        # Estilos generales
        style_commands += [
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, 1), (-1, 1), colors.whitesmoke),
            ('BACKGROUND', (0, 2), (-1, 2), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 2), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 1), (-1, 2), 4),
        ]
        table.setStyle(TableStyle(style_commands))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="tabla_{study_id}.pdf"'
        return response

    except Study.DoesNotExist:
        return HttpResponse(status=404, content="Estudio no encontrado.")
    except Exception as e:
        return HttpResponse(status=500, content=f"Error al generar PDF: {e}")

def generar_pdf_ficha(request):
    study_id = request.GET.get('study_id')
    person_id = request.GET.get('person_id')
    # 1) Recuperar objetos
    study  = Study.objects.get(id=study_id)
    person = Person.objects.get(id=person_id)

    # 2) Formatear datos básicos
    # datos = {
    #     "nombre": person.name,
    #     "sexo":   dict(Person.GENDER_CHOICES).get(person.gender, person.gender),
    #     "pais":   person.country,
    #     "estado": person.state,
    #     "fecha_nacimiento": if  person.date_of_birth:  person.date_of_birth.strftime("%Y/%m/%d") ,
    # }

    datos = {
        "nombre": person.name if person.name else "",  # o '' si prefieres string vacío
        "sexo": dict(Person.GENDER_CHOICES).get(person.gender, person.gender) if person.gender else "",
        "pais": person.country if person.country else "",
        "estado": person.state if person.state else "",
        "fecha_nacimiento": person.date_of_birth.strftime("%Y/%m/%d") if person.date_of_birth else ""
    }

    # 3) Traer todas las mediciones de este estudio y persona
    medidas = Measurement.objects.filter(person=person, study=study)
   
    for m in medidas:
        print(m.position)

    # 4) Separar erecto (P) y sentado (S)
    medidas_erecto  = [
        {"nombre": m.dimension.name, "valor": m.value}
        for m in medidas if m.position == "P"
    ]
    print("Erectp",medidas_erecto)
    medidas_sentado = [
        {"nombre": m.dimension.name, "valor": m.value}
        for m in medidas if m.position == "S"
    ]

    # 5) Tomar peso (dimensión "Peso") si existe
    peso_med = medidas.filter(dimension__name__iexact="Peso").first()
    peso = peso_med.value if peso_med else ""

    # 6) Control: fechas y responsables
    datos["medidas_erecto"]  = medidas_erecto
    datos["medidas_sentado"] = medidas_sentado
    datos["control"] = {
        "fecha":      timezone.now().date().strftime("%Y/%m/%d"),
        "inicio":     study.start_date.strftime("%Y/%m/%d"),
        "fin":        study.end_date.strftime("%Y/%m/%d"),
        "responsable": getattr(study, "responsible", ""),
        "supervisor":  getattr(study, "supervisor", ""),
        "dimensiones": "mm",
        "peso":        peso,
    }

    # 7) Generar PDF
    buffer = io.BytesIO()
    doc     = SimpleDocTemplate(buffer, pagesize=A4)
    styles  = getSampleStyleSheet()
    styleN  = styles["Normal"]
    styleB  = styles["Heading4"]

    def p(texto, estilo=styleN):
        return Paragraph(texto, estilo)

    # Construcción de filas
    tabla = []
    tabla.append([p("<b>Ficha de recolección de datos antropométricos</b>"), "", ""])
    tabla.append([p(f"Nombre: {datos['nombre']}"), p(f"Sexo: {datos['sexo']}"), ""])
    tabla.append([
        p(f"País: {datos['pais']}"),
        p(f"Estado: {datos['estado']}"),
        p(f"Fecha de nacimiento: {datos['fecha_nacimiento']}"),
    ])

    # Erecto
    tabla.append([p("<b>Medidas en posición erecto</b>"), "", ""])
    tabla.append([p("<b>No</b>"), p("<b>Nombre</b>"), p("<b>Valor (mm)</b>")])
    for i, m in enumerate(datos["medidas_erecto"], start=1):
        tabla.append([str(i), m["nombre"], str(m["valor"])])

    # Sentado
    idx_sent =  5 + len(datos["medidas_erecto"])
    tabla.append([p("<b>Medidas en posición sentado</b>"), "", ""])
    tabla.append([p("<b>No</b>"), p("<b>Nombre</b>"), p("<b>Valor (mm)</b>")])
    for i, m in enumerate(datos["medidas_sentado"], start=1):
        tabla.append([str(i), m["nombre"], str(m["valor"])])

    # Control
    idx_ctrl = idx_sent + 2 + len(datos["medidas_sentado"])
    tabla.append([p("<b>Control</b>"), "", ""])
    tabla.append([
        p(f"Fecha: {datos['control']['fecha']}"),
        p(f"Inicio: {datos['control']['inicio']}"),
        p(f"Fin: {datos['control']['fin']}"),
    ])
    tabla.append([
        p(f"Responsable: {datos['control']['responsable']}"),
        p(f"Supervisor: {datos['control']['supervisor']}"),
        "",
    ])
    tabla.append([
        p(f"Dimensiones: {datos['control']['dimensiones']}"),
        p(f"Peso: {datos['control']['peso']} kg"),
        "",
    ])

    # Crear tabla y estilo
    t = Table(tabla, colWidths=[180, 180, 180])
    estilo = TableStyle([
        ('GRID',    (0,0), (-1,-1), 0.8, colors.black),
        ('SPAN',    (0,0), (-1,0)),   # Título
        ('SPAN',    (0,3), (-1,3)),   # Subtítulo erecto
        ('SPAN',    (0, idx_sent), (-1, idx_sent)),  # Subtítulo sentado
        ('SPAN',    (0, idx_ctrl),   (-1, idx_ctrl)),# Subtítulo control
        ('ALIGN',   (0,0), (-1,0),   'CENTER'),
        ('VALIGN',  (0,0), (-1,-1),  'MIDDLE'),
    ])
    t.setStyle(estilo)

    doc.build([t])
    buffer.seek(0)

    return HttpResponse(
        buffer,
        content_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="ficha_{person.name}.pdf"'}
    )


#Funciones auxiliares para calcular tablas antropometricas
def calculate_percentiles_for_study(
    study_id: int,
    gender: str = None,                  # 'M', 'F' o 'mixto'
    age_ranges: list[tuple[int,int]] = None,
    dimensions_filter: list[int] = None,
    percentiles_list: list[float] = None
) -> list:
    """Calcula media, desviación y percentiles por dimensión,
       por género y por cada rango de edad."""
    # Obtener el estudio o lanzar 404
    try:
        study = Study.objects.get(id=study_id)
    except Study.DoesNotExist:
        raise Http404(f"Study with id {study_id} not found")

    # Percentiles por defecto
    percentiles = percentiles_list or [5, 10, 25, 50, 75, 90, 95]

    # Filtrar dimensiones
    dims_qs = Dimension.objects.filter(dimension_study__id_study=study)
    if dimensions_filter:
        dims_qs = dims_qs.filter(id__in=dimensions_filter)

    # Determinar géneros a procesar
    if gender == 'mixto':
        genders_to_process = ['M', 'F']
    elif gender in ('M', 'F'):
        genders_to_process = [gender]
    else:
        genders_to_process = [None]  # todos juntos si no especifica

    results = []

    for dim in dims_qs:
        entry = {
            'dimension': dim.name,
            'dimension_id': dim.id,
            'by_gender': {}
        }

        for g in genders_to_process:
            # Preparamos buckets de edad
            buckets = age_ranges or [(None, None)]
            age_buckets_stats = {}

            for amin, amax in buckets:
                vals = _collect_measurements(
                    study=study,
                    dimension=dim,
                    gender_filter=g,
                    age_min=amin,
                    age_max=amax
                )
                if not vals:
                    continue
                stats = _compute_stats(vals, percentiles)
                key = f"{amin}-{amax}" if amin is not None else "all"
                age_buckets_stats[key] = stats

            if age_buckets_stats:
                age_key = g or 'all'
                entry['by_gender'][age_key] = age_buckets_stats

        if entry['by_gender']:
            results.append(entry)

    return results

def _collect_measurements(study, dimension, gender_filter, age_min, age_max):
    """
    Devuelve lista de valores filtrados según:
      - género (M/F/None),
      - rango de edad [age_min, age_max] (si se pasan),
      - pertenecen al estudio y dimensión dados.
    """
    qs = Measurement.objects.filter(
        study=study,
        dimension=dimension
    ).select_related('person')

    vals = []
    for m in qs:
        p = m.person
        # Filtrar por género
        if gender_filter and p.gender != gender_filter:
            continue

        # Filtrar por rango de edad
        if (age_min is not None or age_max is not None) and p.date_of_birth and m.date:
            md = m.date.date()
            bd = p.date_of_birth
            age = md.year - bd.year - ((md.month, md.day) < (bd.month, bd.day))
            if age_min is not None and age < age_min:
                continue
            if age_max is not None and age > age_max:
                continue

        vals.append(m.value)

    return vals

def _compute_stats(values, percentiles_list):
    """
    Calcula media, desviación estándar y percentiles de la lista de valores.
    """
    arr = np.array(values, dtype=float)
    mean = float(arr.mean())
    sd = float(arr.std(ddof=0))
    perc_vals = np.percentile(arr, percentiles_list)
    perc_dict = {str(int(p)): float(v) for p, v in zip(percentiles_list, perc_vals)}

    return {
        'mean': mean,
        'sd': sd,
        'percentiles': perc_dict
    }
